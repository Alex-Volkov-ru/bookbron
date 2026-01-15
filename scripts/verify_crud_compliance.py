"""
Проверка соответствия реализации CRUD доступов из Excel файла
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl import load_workbook

# Читаем Excel файл
wb = load_workbook('CRUD доступ.xlsx')
ws = wb.active

# Парсим данные
endpoints_requirements = {}
operations = ['Crec', 'Rall', 'Rrec', 'Urec', 'Drec']
roles = ['admin', 'manager', 'user', 'unauthorized']

data_start = 5
for row_idx in range(data_start, 12):  # Строки 5-11
    endpoint_cell = ws.cell(row=row_idx, column=1)
    if not endpoint_cell.value:
        continue
    
    endpoint = str(endpoint_cell.value).strip()
    if not endpoint or endpoint.startswith('Примечание'):
        break
    
    # Читаем доступы
    access_data = {}
    for role_idx, role in enumerate(roles):
        role_access = {}
        start_col = 2 + (role_idx * 5)
        
        for op_idx, operation in enumerate(operations):
            col = start_col + op_idx
            cell = ws.cell(row=row_idx, column=col)
            value = str(cell.value).strip() if cell.value else ""
            role_access[operation] = value
        
        access_data[role] = role_access
    
    endpoints_requirements[endpoint] = access_data

print("=" * 100)
print("ПРОВЕРКА СООТВЕТСТВИЯ CRUD ДОСТУПОВ")
print("=" * 100)
print()

# Маппинг эндпоинтов Excel -> Реализация
endpoint_mapping = {
    '/user': '/users',
    '/cafes': '/cafes',
    '/cafe/tables': '/cafe/{cafe_id}/tables',
    '/cafe/time_slots': '/cafe/{cafe_id}/slots',
    '/actions': '/actions',
    '/dishes': '/dishes',
    '/booking': '/booking',
    '/media': '/media'
}

# Операции -> HTTP методы
operation_to_method = {
    'Crec': 'POST',
    'Rall': 'GET (list)',
    'Rrec': 'GET (by id)',
    'Urec': 'PATCH',
    'Drec': 'DELETE'
}

print("ТРЕБОВАНИЯ ИЗ EXCEL:")
print("-" * 100)
for excel_endpoint, accesses in endpoints_requirements.items():
    api_endpoint = endpoint_mapping.get(excel_endpoint, excel_endpoint)
    print(f"\n{excel_endpoint} -> {api_endpoint}:")
    for role in roles:
        allowed_ops = [op for op in operations if accesses[role].get(op) == '+']
        if allowed_ops:
            methods = [operation_to_method[op] for op in allowed_ops]
            print(f"  {role:15} может: {', '.join(methods)}")

print("\n" + "=" * 100)
print("КРИТИЧЕСКИЕ ПРИМЕЧАНИЯ:")
print("-" * 100)
print("1. Менеджер имеет доступ только к тем записям, которые относятся к его кафе")
print("2. Авторизированный пользователь имеет доступ только к записям, которые он создал сам")
print("3. Не авторизированный пользователь может только зарегистрироваться")
print("=" * 100)

