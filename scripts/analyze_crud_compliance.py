"""
Анализ соответствия CRUD доступов из Excel файла с реализацией
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl import load_workbook

# Читаем Excel файл
wb = load_workbook('CRUD доступ.xlsx')
ws = wb.active

# Парсим данные
endpoints = {}
headers_row = 3
data_start = 5

# Читаем заголовки ролей
role_headers = []
for col_idx in range(2, ws.max_column + 1):
    cell = ws.cell(row=headers_row, column=col_idx)
    if cell.value:
        role_headers.append(str(cell.value).strip())

# Читаем операции (Crec, Rall, Rrec, Urec, Drec)
operations = []
for col_idx in range(2, 7):  # Колонки 2-6 для админа
    cell = ws.cell(row=data_start-1, column=col_idx)
    if cell.value:
        operations.append(str(cell.value).strip())

print("=" * 100)
print("АНАЛИЗ СООТВЕТСТВИЯ CRUD ДОСТУПОВ")
print("=" * 100)
print(f"\nОперации: {operations}")
print(f"Роли: Администратор, Менеджер, Пользователь (авторизированный), Пользователь (не авторизированный)")
print()

# Читаем данные по эндпоинтам
for row_idx in range(data_start, ws.max_row + 1):
    endpoint_cell = ws.cell(row=row_idx, column=1)
    if not endpoint_cell.value:
        continue
    
    endpoint = str(endpoint_cell.value).strip()
    if not endpoint or endpoint.startswith('Примечание') or endpoint.startswith('Пояснения'):
        break
    
    # Читаем доступы для каждой роли
    access_data = {}
    role_idx = 0
    for col_group in range(0, 4):  # 4 роли
        role_access = {}
        start_col = 2 + (col_group * 5)  # Каждая роль занимает 5 колонок
        
        for op_idx, operation in enumerate(operations):
            col = start_col + op_idx
            cell = ws.cell(row=row_idx, column=col)
            value = str(cell.value).strip() if cell.value else ""
            role_access[operation] = value
        
        if col_group == 0:
            access_data['admin'] = role_access
        elif col_group == 1:
            access_data['manager'] = role_access
        elif col_group == 2:
            access_data['user'] = role_access
        elif col_group == 3:
            access_data['unauthorized'] = role_access
    
    endpoints[endpoint] = access_data

print("ЭНДПОИНТЫ И ДОСТУПЫ:")
print("-" * 100)
for endpoint, accesses in endpoints.items():
    print(f"\n{endpoint}:")
    for role, ops in accesses.items():
        allowed = [op for op, val in ops.items() if val == '+']
        denied = [op for op, val in ops.items() if val == '-']
        print(f"  {role:15} | Разрешено: {allowed} | Запрещено: {denied}")

print("\n" + "=" * 100)
print("ПРИМЕЧАНИЯ ИЗ EXCEL:")
print("-" * 100)
for row_idx in range(14, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=1)
    if cell.value and ('Примечание' in str(cell.value) or 'Пояснения' in str(cell.value) or 
                       'Менеджер' in str(cell.value) or 'Авторизированный' in str(cell.value) or
                       'Не авторизированный' in str(cell.value) or 'доступ разрешен' in str(cell.value)):
        print(f"  {cell.value}")

print("\n" + "=" * 100)

