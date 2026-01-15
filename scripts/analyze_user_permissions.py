"""
Детальный анализ прав пользователей из всех документов
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl import load_workbook

# Читаем Excel файл
wb = load_workbook('CRUD доступ.xlsx')
ws = wb.active

print("=" * 100)
print("АНАЛИЗ ПРАВ ПОЛЬЗОВАТЕЛЕЙ ИЗ CRUD доступ.xlsx")
print("=" * 100)
print()

# Парсим данные
endpoints_requirements = {}
operations = ['Crec', 'Rall', 'Rrec', 'Urec', 'Drec']
roles = ['admin', 'manager', 'user', 'unauthorized']

data_start = 5
for row_idx in range(data_start, 12):  # Строки 5-11
    endpoint_cell = ws.cell(row=row_idx, column=1)
    if not endpoint_cell.value:
        continue
    
    endpoint = str(endpoint_cell.value).strip()
    if not endpoint or endpoint.startswith('Примечание'):
        break
    
    # Читаем доступы
    access_data = {}
    for role_idx, role in enumerate(roles):
        role_access = {}
        start_col = 2 + (role_idx * 5)
        
        for op_idx, operation in enumerate(operations):
            col = start_col + op_idx
            cell = ws.cell(row=row_idx, column=col)
            value = str(cell.value).strip() if cell.value else ""
            role_access[operation] = value
        
        access_data[role] = role_access
    
    endpoints_requirements[endpoint] = access_data

print("ВИДЫ ПОЛЬЗОВАТЕЛЕЙ:")
print("-" * 100)
print("1. Администратор (admin) - полный доступ")
print("2. Менеджер (manager) - доступ только к своим кафе")
print("3. Пользователь авторизированный (user) - доступ только к своим записям")
print("4. Пользователь не авторизированный (unauthorized) - только регистрация")
print()

print("ПРАВА ПО ЭНДПОИНТАМ:")
print("-" * 100)
for endpoint, accesses in endpoints_requirements.items():
    print(f"\n{endpoint}:")
    for role in roles:
        allowed_ops = [op for op in operations if accesses[role].get(op) == '+']
        denied_ops = [op for op in operations if accesses[role].get(op) == '-']
        if allowed_ops or denied_ops:
            print(f"  {role:15} | Разрешено: {allowed_ops} | Запрещено: {denied_ops}")

print("\n" + "=" * 100)
print("КРИТИЧЕСКИЕ ПРИМЕЧАНИЯ ИЗ EXCEL:")
print("-" * 100)
for row_idx in range(14, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=1)
    if cell.value:
        value = str(cell.value).strip()
        if 'Примечание' in value or 'Менеджер' in value or 'Авторизированный' in value or 'Не авторизированный' in value:
            print(f"  {value}")

print("\n" + "=" * 100)
print("ВЫВОДЫ:")
print("-" * 100)
print("1. Неавторизированный пользователь может ТОЛЬКО зарегистрироваться (POST /users)")
print("2. НО! По логике, неавторизированные должны видеть кафе для выбора перед регистрацией")
print("3. Авторизированный пользователь может управлять только своими записями")
print("4. Менеджер может управлять только своими кафе")
print("5. Админ имеет полный доступ")
print("=" * 100)

