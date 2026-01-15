"""
Скрипт для проверки соответствия CRUD доступов из Excel файла
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

try:
    import openpyxl
    from openpyxl import load_workbook
    
    # Читаем Excel файл
    wb = load_workbook('CRUD доступ.xlsx')
    ws = wb.active
    
    print("=" * 80)
    print("CRUD ДОСТУПЫ ИЗ EXCEL ФАЙЛА")
    print("=" * 80)
    print(f"Лист: {ws.title}")
    print(f"Строк: {ws.max_row}, Колонок: {ws.max_column}")
    print()
    
    # Читаем заголовки
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    print("Заголовки:", headers)
    print()
    
    # Читаем данные
    print("Данные:")
    print("-" * 80)
    for row_idx in range(2, min(ws.max_row + 1, 100)):  # Первые 100 строк
        row_data = []
        for cell in ws[row_idx]:
            if cell.value:
                row_data.append(str(cell.value).strip())
            else:
                row_data.append("")
        
        if any(row_data):  # Пропускаем пустые строки
            print(f"Строка {row_idx}: {row_data}")
    
    print()
    print("=" * 80)
    
except ImportError:
    print("Ошибка: openpyxl не установлен")
    print("Установите: pip install openpyxl")
    print()
    print("Попытка прочитать как CSV или текстовый файл...")
    
    # Попробуем прочитать как текстовый файл
    try:
        with open('CRUD доступ.xlsx', 'rb') as f:
            content = f.read(1000)
            print("Файл найден, но это бинарный Excel файл")
            print("Нужна библиотека openpyxl для чтения")
    except Exception as e:
        print(f"Ошибка: {e}")

